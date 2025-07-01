import argparse
import os
from collections import defaultdict
from pathlib import Path

import openpyxl
from draw_packet import make_table, write_packets

parser = argparse.ArgumentParser()
parser.add_argument(
    '--product-type',
    type=str,
    default="client_value",
    help='product type: client_value, client, mobile')
args = parser.parse_args()

total_pif_num = 0
fail_count = 0

parent_directory_path = os.path.dirname(os.path.abspath(
    os.path.dirname(os.path.abspath(
        os.path.dirname(os.path.abspath(__file__))))))

product_type = args.product_type

pif_dir = 'provided_interface'
target_dir = [
    os.path.join(
        parent_directory_path,
        'core',
        pif_dir),
    os.path.join(
        parent_directory_path,
        'product',
        product_type,
        pif_dir)]

header_file = 'core_pif.py'
header_dir = os.path.join(
    parent_directory_path,
    'core',
    'framework',
    header_file)
header_dict = defaultdict(list)

excel_filename = 'packets_' + product_type + '.xlsx'
output_folder = os.path.join(parent_directory_path, 'output', 'packetIF')
Path(output_folder).mkdir(parents=True, exist_ok=True)
dest_filename = os.path.join(output_folder, excel_filename)
except_file_list = ['nand_pif']


work_sheet_name = []
is_make_table_sheet = True


class IndexErrorException(Exception):
    def __init__(self, value):
        super().__init__(f'\t!! Error: Invalid DWORD Size ({value})')


class ValueErrorException(Exception):
    def __init__(self, value):
        super().__init__(f'\t!! Error: Invalid DWORD Size ({value})')


class InvalidParsingException(Exception):
    def __init__(self, value):
        super().__init__(f'\t!! Error: Invalid Parsing : ({value})')


def get_sq_name(line):
    try:
        return line.split('class')[1].split('(')[0].strip()
    except BaseException:
        raise InvalidParsingException(line.strip())


def get_field(line):
    try:
        return line.split(',')[0].split("'")[1].strip()
    except BaseException:
        raise InvalidParsingException(line.strip())


def get_dw(line):
    try:
        return int(line.split(',')[1].split('#')[1].split('dw')[1].strip())
    except BaseException:
        raise InvalidParsingException(line.strip())


def get_bit(line):
    try:
        start_bit = int(line.split(',')[2].split('_')[0].strip())
        try:
            end_bit = int(line.split(',')[2].split('_')[1].strip())
        except BaseException:
            end_bit = start_bit
        return start_bit, end_bit
    except BaseException:
        raise InvalidParsingException(line.strip())


def get_description(line):
    try:
        return line.split(',')[3].strip()
    except BaseException:
        raise InvalidParsingException(line.strip())


def parse_header_dict():
    with open(header_dir, 'r', encoding='UTF-8') as f:
        while line := f.readline():
            if 'head_key_list' in line:
                while line := f.readline():
                    if ']' in line:
                        return
                    if '\n' == line:
                        continue
                    header_dict[get_dw(line)].append({'field': get_field(
                        line), 'bit': get_bit(line), 'desc': get_description(line)})


def parse_packet_file(file_path):
    module = {}
    with open(file_path, 'r', encoding='UTF-8') as f:
        while line := f.readline():
            if 'class' in line:
                sq_name = get_sq_name(line)
                module[sq_name] = header_dict.copy()
            elif ('body_key_list' in line) and ('=' in line):
                while line := f.readline():
                    if ']' in line:
                        break
                    if '\n' == line:
                        continue
                    module[sq_name][get_dw(line)].append({'field': get_field(
                        line), 'bit': get_bit(line), 'desc': get_description(line)})
    return module


def get_pif_name(file_path):
    return os.path.splitext(os.path.split(file_path)[-1])[0]


def get_pif_files():
    filepath_list = sorted([os.path.join(cur_dir, name)
                            for cur_dir in target_dir for name in os.listdir(cur_dir)])
    filename_dict = {get_pif_name(
        file_path): file_path for file_path in filepath_list if file_path.endswith('_pif.py')}
    for except_file in except_file_list:
        try:
            filename_dict.pop(except_file)
            print(f'Except PIF File : {except_file}')
        except KeyError:
            print(f'Fail Except PIF File name : {except_file} ')
    return filename_dict


if __name__ == "__main__":
    packet_file_dict = get_pif_files()
    print(
        f"Translating PIF files into xlsx format (Total PIF files: {len(packet_file_dict)})")

    work_book = openpyxl.Workbook()
    work_sheet = None
    total_pif_num = len(packet_file_dict)

    try:
        parse_header_dict()
    except InvalidParsingException as err:
        print(f'{err}')

    for file_name, file_path in packet_file_dict.items():
        print(' -', file_name)
        try:
            packet = parse_packet_file(file_path)
            work_sheet = work_book.active if work_sheet is None else work_book.create_sheet()
            work_sheet.title = file_name
            write_packets(work_sheet, packet)
            work_sheet_name.append(work_sheet.title)
        except IndexError as index_error:
            print(f'\t!! Index Error {index_error}')
            fail_count += 1
        except ValueError as value_error:
            print(f'\t!! ValueError {value_error}')
            fail_count += 1
        except InvalidParsingException as err:
            print(f'{err}')
            fail_count += 1

    if is_make_table_sheet:
        work_sheet = work_book.active if work_sheet is None else work_book.create_sheet(
            "Table", 0)
        make_table(work_sheet, work_sheet_name)
    work_book.save(dest_filename)

    print("Translate Result : {} / {} (Success / Total)".format(total_pif_num -
                                                                fail_count, total_pif_num))
