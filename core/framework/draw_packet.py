import re
from dataclasses import dataclass

from openpyxl.styles import (Alignment, Border, Color, Font, NamedStyle,
                             PatternFill, Side)


@dataclass
class Position:
    start_row: int
    start_column: int
    row_width: int = 0
    column_width: int = 0

    def set_start_row(self, val):
        self.start_row = val

    def set_start_col(self, val):
        self.start_column = val

    def move_start_row(self, val):
        self.start_row += val

    def move_start_col(self, val):
        self.start_column += val

    def move_start_dword(self):
        self.start_row += 1
        self.start_column += WORD_SIZE

    def move_next_dword(self):
        self.start_row += 2

    def move_start_line(self):
        self.start_row -= 1
        self.start_column += WORD_SIZE

    def move_back_start_line(self):
        self.start_row -= 1
        self.start_column -= WORD_SIZE

    def __repr__(self):
        return f'({self.start_row},{self.start_column}) ~ ({self.start_row + self.row_width},{self.start_column + self.column_width})'


thin_side = Side('thin', color='000000')
thin_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side)
header_fill = PatternFill(patternType='solid', fgColor=Color('ddd9c4'))
rsvd_fill = PatternFill(patternType='solid', fgColor=Color('d9d9d9'))
desc_fill = PatternFill(patternType='solid', fgColor=Color('e8e8e6'))
header_font = Font(name='Calibri', bold=True, color='000000', size=11)
cell_font = Font(name='Calibri', color='000000', size=10)
desc_header_font = Font(
    name='Calibri',
    italic=True,
    bold=True,
    color='000000',
    size=10)
desc_font = Font(name='Calibri', italic=True, color='000000', size=10)
center_align = Alignment(
    vertical='center',
    horizontal='center',
    wrap_text=True)
header_style = NamedStyle(
    name='header',
    font=header_font,
    border=thin_border,
    alignment=center_align,
    fill=header_fill)
field_style = NamedStyle(
    name='field',
    font=cell_font,
    border=thin_border,
    alignment=center_align)
desc_header_style = NamedStyle(
    name='desc_header',
    font=desc_header_font,
    border=thin_border,
    alignment=center_align,
    fill=desc_fill)
desc_field_style = NamedStyle(
    name='desc_field',
    font=desc_font,
    border=thin_border,
    alignment=center_align)

DWORD_SIZE = 32
WORD_SIZE = 16

max_bit_size = DWORD_SIZE
max_field_row_size = WORD_SIZE

dword_row = 2

sheet_current_row = 0
sheet_current_column = 0


def merge_cell_and_write(work_sheet, position, val, style):
    end_row = position.start_row + position.row_width
    end_column = position.start_column + position.column_width
    work_sheet.merge_cells(
        start_row=position.start_row,
        start_column=position.start_column,
        end_row=end_row,
        end_column=end_column)

    for row in range(position.start_row, end_row + 1):
        for col in range(position.start_column, end_column + 1):
            work_sheet.cell(row=row, column=col).border = style.border

    top_left_cell = work_sheet.cell(
        row=position.start_row,
        column=position.start_column)
    top_left_cell.value = val
    top_left_cell.style = style

    if re.findall('rsvd', val.lower()):
        top_left_cell.fill = rsvd_fill


def apply_cell_style(cell, val, style):
    cell.value = val
    cell.style = style


def make_table(ws, work_book_name_list):
    sheet_current_column = 2
    sheet_current_row = 2

    apply_cell_style(
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column),
        'Module',
        header_style)
    apply_cell_style(
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column +
            1),
        'PIF',
        header_style)

    sheet_current_row += 1
    max_column_length = 0
    for work_book_name in work_book_name_list:
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column).value = work_book_name.strip().split('_pif')[0]
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column).border = thin_border
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column).alignment = center_align
        max_column_length = max(len(work_book_name), max_column_length)
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column +
            1).value = work_book_name
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column +
            1).style = "Hyperlink"
        ws.cell(row=sheet_current_row, column=sheet_current_column +
                1).hyperlink = "#" + work_book_name + "!A1"
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column +
            1).border = thin_border
        ws.cell(
            row=sheet_current_row,
            column=sheet_current_column +
            1).alignment = center_align

        sheet_current_row += 1
    ws.column_dimensions["B"].width = max_column_length * 2
    ws.column_dimensions["C"].width = max_column_length * 2


def make_packet_format_header(work_sheet):
    global sheet_current_row, sheet_current_column

    header_row_width = 2
    header_column_width = 0
    bit_row_width = 0
    bit_column_width = 0
    header_name = ['Index', 'Arg Type', 'DW/Bit']

    work_sheet.column_dimensions['A'].width = 10
    work_sheet.column_dimensions['B'].width = 20
    work_sheet.column_dimensions['C'].width = 8

    sr = 1
    sc = 1

    for idx, name in enumerate(header_name):
        merge_cell_and_write(
            work_sheet,
            Position(
                sr,
                sc,
                header_row_width,
                header_column_width),
            name,
            header_style)
        sc += 1

    for bit in range(max_bit_size):
        value = str(max_bit_size - (bit + 1))
        merge_cell_and_write(work_sheet,
                             Position(sr + (bit // max_field_row_size),
                                      sc + (bit % max_field_row_size),
                                      bit_row_width,
                                      bit_column_width),
                             value,
                             header_style)

    sr += (max_bit_size // max_field_row_size)

    merge_cell_and_write(
        work_sheet,
        Position(
            sr,
            sc,
            bit_row_width,
            (max_field_row_size - 1)),
        'Message',
        header_style)

    sheet_current_row = sr + 1
    sheet_current_column = 1


def is_width_dword(width):
    return width == DWORD_SIZE


def is_bigger_line_word(total_line_bit):
    return total_line_bit > WORD_SIZE


def write_packets(work_sheet, packets):
    make_packet_format_header(work_sheet)
    for idx, packet_name in enumerate(packets.keys()):
        write_packet(work_sheet, packets[packet_name], packet_name, idx)


def write_packet(work_sheet, packet, packet_name, packet_idx):
    global sheet_current_row, sheet_current_column

    total_dw_count = len(packet.values())
    make_sq_header(work_sheet, total_dw_count, packet_name, packet_idx)
    write_field(work_sheet, packet, total_dw_count)

    make_description_header(work_sheet)
    write_description(work_sheet, packet)


def make_sq_header(work_sheet, total_dw_count, packet_name, packet_idx):
    global sheet_current_row, sheet_current_column

    wp = Position(sheet_current_row, sheet_current_column)
    packet_row_height = (total_dw_count * dword_row)

    next_col = 1

    for name in [str(packet_idx), packet_name]:
        merge_cell_and_write(
            work_sheet,
            Position(
                wp.start_row,
                wp.start_column,
                packet_row_height - 1,
                0),
            name,
            field_style)
        wp.move_start_col(next_col)

    for dw in range(total_dw_count):
        merge_cell_and_write(
            work_sheet,
            Position(
                wp.start_row,
                wp.start_column,
                dword_row - 1,
                0),
            str(dw),
            field_style)
        wp.move_start_row(dword_row)

    sheet_current_column = wp.start_column


def write_field(work_sheet, packet, total_dw_count):
    global sheet_current_row, sheet_current_column
    packet_row_height = total_dw_count * dword_row
    wp = Position(sheet_current_row, sheet_current_column)

    for dw, fields in packet.items():
        wp.move_start_dword()
        total_line_bit = 0
        odd_in_dword = True

        for field in fields:
            start_bit, end_bit = field['Bit']
            size_bit = end_bit - start_bit + 1
            name = field['Field']
            total_line_bit += size_bit

            if is_width_dword(size_bit):
                wp.move_back_start_line()
                merge_cell_and_write(
                    work_sheet,
                    Position(
                        wp.start_row,
                        wp.start_column + 1,
                        1,
                        WORD_SIZE - 1),
                    name,
                    field_style)
            elif is_bigger_line_word(total_line_bit):
                spare_bit = total_line_bit - WORD_SIZE
                size_bit = WORD_SIZE - (total_line_bit - size_bit)
                for bit in [size_bit, spare_bit]:
                    wp.move_start_col(-1 * bit)
                    merge_cell_and_write(
                        work_sheet,
                        Position(
                            wp.start_row,
                            wp.start_column + 1,
                            0,
                            bit - 1),
                        name,
                        field_style)
                    if total_line_bit > WORD_SIZE:
                        wp.move_start_line()
                        total_line_bit -= WORD_SIZE
                        odd_in_dword = False
            else:
                wp.move_start_col(-1 * size_bit)
                merge_cell_and_write(
                    work_sheet,
                    Position(
                        wp.start_row,
                        wp.start_column + 1,
                        0,
                        size_bit - 1),
                    name,
                    field_style)
                if total_line_bit == WORD_SIZE and odd_in_dword:
                    wp.move_start_line()
                    odd_in_dword = False
                    total_line_bit = 0
        wp.move_next_dword()

    sheet_current_row += packet_row_height
    sheet_current_column = 1


def make_description_header(work_sheet):
    global sheet_current_row, sheet_current_column

    header_row_height = 0
    header_column_width = 0
    header_name = ['Field Name', 'Width']

    sr = sheet_current_row
    sc = sheet_current_column + 1

    for idx, name in enumerate(header_name):
        merge_cell_and_write(
            work_sheet,
            Position(
                sr,
                sc,
                header_row_height,
                header_column_width),
            name,
            desc_header_style)
        sc += 1

    merge_cell_and_write(
        work_sheet,
        Position(
            sr,
            sc,
            0,
            (max_field_row_size - 1)),
        'Description',
        desc_header_style)

    sheet_current_row = sr + 1
    sheet_current_column = 1


def write_description(work_sheet, packet):
    global sheet_current_row, sheet_current_column

    sr = sheet_current_row
    sc = sheet_current_column + 1

    for _, fields in packet.items():
        for field in fields:
            name = field['field']
            if re.findall('rsvd', name.lower()):
                continue
            start_bit, end_bit = field['bit']
            size_bit = end_bit - start_bit + 1

            for text, col_width in [
                    (name, 0), (str(size_bit), 0), (field['desc'], max_field_row_size - 1)]:
                merge_cell_and_write(work_sheet,
                                     Position(sr, sc, 0, col_width),
                                     text, desc_field_style)
                sc += 1

            sr += 1
            sc = sheet_current_column + 1

    sheet_current_row = sr + 1
    sheet_current_column = 1
